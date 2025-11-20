import streamlit as st
import openpyxl
import pandas as pd
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

# --- Constants ---
BGN_TO_EUR_DIVISOR = Decimal('1.95583')
TWO_PLACES = Decimal('0.01')

# --- Helper Functions ---

def parse_decimal(number_str):
    """Takes a price string/float/int, cleans it, and returns a Decimal."""
    if number_str is None: return None
    number_str = str(number_str)
    try:
        cleaned_str = number_str.replace(' ', '').replace(',', '.')
        return Decimal(cleaned_str)
    except InvalidOperation:
        return None

def normalize(decimal):
    normalized = decimal.quantize(TWO_PLACES, rounding=ROUND_HALF_UP)
    return normalized

def get_column_index(sheet, column_name):
    """Finds the 1-based index of a column header."""
    target = column_name.strip()
    for cell in sheet[1]: 
        if cell.value and str(cell.value).strip() == target:
            return cell.column
    return None

def verify_row(bgn_val, eur_val):
    """
    Calculates and compares values. 
    Returns a tuple: (is_mismatch, calculated_eur, normalized_eur_given)
    """
    try:
        if bgn_val is None:
            return False, None, None

        calculated_eur = bgn_val / BGN_TO_EUR_DIVISOR
        normalized_calc_eur = normalize(calculated_eur)
        
        if eur_val is not None:
            eur_price_normalized = normalize(eur_val)
            
            if normalized_calc_eur != eur_price_normalized:
                return True, normalized_calc_eur, eur_price_normalized
            
        return False, normalized_calc_eur, None

    except Exception:
        return False, None, None

# --- Main Streamlit App ---

def main():
    st.set_page_config(page_title="BGN/EUR Verifier", layout="wide")
    
    st.title("BGN to EUR Excel Tables Verifier")
    st.markdown(f"""
    Upload two Excel files to verify currency conversions.
    **Conversion Rate:** $1 \\text{{ EUR}} = {BGN_TO_EUR_DIVISOR} \\text{{ BGN}}$
    """)

    col1, col2 = st.columns(2)
    
    with col1:
        bgn_file = st.file_uploader("Upload BGN Source File", type=['xlsx'])
    with col2:
        eur_file = st.file_uploader("Upload EUR Target File", type=['xlsx'])

    if bgn_file and eur_file:
        st.divider()
        
        try:
            # Load Workbooks
            # We load data_only=True so we get values, not formulas
            bgn_wb = openpyxl.load_workbook(bgn_file, data_only=True)
            eur_wb = openpyxl.load_workbook(eur_file, data_only=True)
            
            bgn_ws = bgn_wb.active
            eur_ws = eur_wb.active
            
            # Extract headers to let user choose columns
            bgn_headers = [str(cell.value).strip() for cell in bgn_ws[1] if cell.value]
            eur_headers = [str(cell.value).strip() for cell in eur_ws[1] if cell.value]
            
            # Find common headers
            common_headers = list(set(bgn_headers).intersection(eur_headers))
            
            st.subheader("Configuration")
            
            selected_columns = st.multiselect(
                "Select Columns to Verify",
                options=bgn_headers,
                default=common_headers,
                help="Select the columns that exist in both files that you want to check."
            )
            
            if st.button("Run Verification", type="primary"):
                if not selected_columns:
                    st.error("Please select at least one column.")
                    return

                # Progress Bar
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                mismatches = []
                column_map = {}
                
                # Map columns
                for col_name in selected_columns:
                    b_idx = get_column_index(bgn_ws, col_name)
                    e_idx = get_column_index(eur_ws, col_name)
                    if b_idx and e_idx:
                        column_map[col_name] = (b_idx, e_idx)
                
                max_row = max(bgn_ws.max_row, eur_ws.max_row)
                rows_to_check = range(2, max_row + 1)
                total_rows = len(rows_to_check)

                # --- Processing Loop ---
                for i, r in enumerate(rows_to_check):
                    # Update progress every 100 rows to save UI renders
                    if i % 100 == 0:
                        progress_bar.progress((i + 1) / total_rows)
                        status_text.text(f"Processing Row {r}...")

                    for col_name, (b_idx, e_idx) in column_map.items():
                        bgn_cell = bgn_ws.cell(row=r, column=b_idx).value
                        eur_cell = eur_ws.cell(row=r, column=e_idx).value
                        
                        bgn_val = parse_decimal(bgn_cell)
                        eur_val = parse_decimal(eur_cell)
                        
                        is_mismatch, calc_eur, given_eur = verify_row(bgn_val, eur_val)
                        
                        if is_mismatch:
                            mismatches.append({
                                "Row": r,
                                "Column": col_name,
                                "Source BGN": f"{bgn_val}",
                                "Calculated EUR": f"{calc_eur}",
                                "File EUR": f"{given_eur}",
                                "Diff": f"{calc_eur - given_eur}"
                            })

                # Finalize UI
                progress_bar.progress(100)
                status_text.text("Done.")
                
                st.divider()
                st.subheader("Results")
                
                if not mismatches:
                    st.success(f"✅ Verification Complete. No mismatches found in {total_rows} rows across {len(selected_columns)} columns.")
                else:
                    st.error(f"❌ Found {len(mismatches)} mismatches.")
                    
                    # Convert list of dicts to DataFrame for nice display
                    df_results = pd.DataFrame(mismatches)
                    st.dataframe(df_results, use_container_width=True)
                    
                    # Add Download Button for CSV
                    csv = df_results.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="Download Mismatch Report (CSV)",
                        data=csv,
                        file_name='mismatch_report.csv',
                        mime='text/csv',
                    )

        except Exception as e:
            st.error(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()